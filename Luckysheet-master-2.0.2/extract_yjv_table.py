import openpyxl
import json
import sys

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

excel_file = '数据库.xlsx'
print(f"读取 {excel_file} 中的YJV电缆数据...\n")

try:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    if 'YJV' in wb.sheetnames:
        ws = wb['YJV']
        
        cables = []
        current_section = None
        
        # 从第3行开始读取（跳过标题）
        for row_idx in range(3, ws.max_row + 1):
            row = ws[row_idx]
            
            # 读取型号规格（第1列）
            spec = row[0].value
            if not spec:
                continue
                
            # 判断是否是新的分类标题
            if 'YJV' in str(spec) and '交联聚乙烯绝缘聚氯乙烯护套电力电缆' in str(spec):
                current_section = str(spec).split()[0]  # 获取电缆类型如 "YJV 0.6/1KV"
                print(f"找到分类: {current_section}")
                continue
            
            # 跳过表头行
            if spec == '型号规格' or spec == 'mm 2':
                continue
            
            try:
                cable = {
                    'section': current_section,
                    'spec': str(spec),
                    'insulation': row[1].value if row[1].value else '',
                    'sheath': row[2].value if row[2].value else '',
                    'approx_outer_diameter': row[3].value if row[3].value else '',
                    'approx_weight': row[4].value if row[4].value else '',
                    'max_conductor_dc_resistance': row[5].value if row[5].value else '',
                    'rated_voltage': row[6].value if row[6].value else '',
                    'test_voltage': row[7].value if row[7].value else '',
                    'min_bending_radius': row[8].value if row[8].value else ''
                }
                
                # 只添加有效数据
                if cable['spec'] and cable['spec'] != 'mm':
                    cables.append(cable)
                    
            except Exception as e:
                print(f"第{row_idx}行解析错误: {e}")
                continue
        
        print(f"\n成功提取 {len(cables)} 条电缆数据")
        
        # 保存为JSON
        with open('dist-refactored/yjv_full_table.json', 'w', encoding='utf-8') as f:
            json.dump(cables, f, ensure_ascii=False, indent=2)
        
        # 生成JavaScript文件
        with open('dist-refactored/yjv_full_table.js', 'w', encoding='utf-8') as f:
            f.write('// YJV电缆完整数据表 - 从 数据库.xlsx 自动生成\n\n')
            f.write('const YJV_FULL_TABLE = ')
            f.write(json.dumps(cables, ensure_ascii=False, indent=2))
            f.write(';\n\n')
            f.write('// 导出数据\n')
            f.write('if (typeof module !== "undefined" && module.exports) {\n')
            f.write('  module.exports = YJV_FULL_TABLE;\n')
            f.write('}\n')
        
        print("已生成文件:")
        print("  - dist-refactored/yjv_full_table.json")
        print("  - dist-refactored/yjv_full_table.js")
        
        # 显示前5条数据示例
        print("\n数据示例（前5条）:")
        for i, cable in enumerate(cables[:5], 1):
            print(f"{i}. {cable['spec']} - {cable['section']}")
        
    else:
        print("错误: 找不到 'YJV' 工作表")
        print("可用的工作表:", wb.sheetnames)
        
except Exception as e:
    print(f'错误: {e}')
    import traceback
    traceback.print_exc()


