# -*- coding: utf-8 -*-
import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('电机电力电气计算表（11月）.xlsx', data_only=True)

ws = wb['风机性能表']

print("风机性能表结构:")
print("=" * 120)
print(f"最大行数: {ws.max_row}, 最大列数: {ws.max_column}\n")

# 显示所有数据
print("完整数据内容:\n")
for row_idx in range(1, min(30, ws.max_row + 1)):
    print(f"第{row_idx}行:")
    for col_idx in range(1, min(20, ws.max_column + 1)):
        val = ws.cell(row_idx, col_idx).value
        if val is not None:
            print(f"  {chr(64+col_idx)}{row_idx}: {val}")
    print()

# 提取所有数据
all_data = {}
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row_idx, col_idx).value
        if val is not None:
            coord = f"{chr(64+col_idx) if col_idx <= 26 else 'A'+ chr(64+col_idx-26)}{row_idx}"
            all_data[coord] = val

# 保存完整数据
with open('fan_performance_full.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"完整数据已保存到 fan_performance_full.json")
print(f"共 {len(all_data)} 个单元格有数据")

wb.close()


