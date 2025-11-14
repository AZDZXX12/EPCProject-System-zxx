# -*- coding: utf-8 -*-
"""
提取风机选型系统所需的完整数据
"""

import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 打开Excel文件
wb = openpyxl.load_workbook('电机电力电气计算表（11月）.xlsx', data_only=True)

print("Excel中的工作表:", wb.sheetnames)

# 找到风机相关的工作表
fan_sheets = [name for name in wb.sheetnames if '风' in name or '扇' in name]
print(f"\n找到风机相关工作表: {fan_sheets}")

# ==================== 1. 提取风机性能表 ====================
print("\n===== 提取风机性能表 =====")
ws_performance = wb.sheetnames[15]  # 索引15是"风机性能表"
print(f"读取工作表: {ws_performance}")
ws = wb[ws_performance]

fan_performance_db = []
headers = []

# 读取表头
for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(1, col_idx).value
    if header:
        headers.append(str(header).strip())

print(f"找到表头: {headers}")

# 读取数据行
for row_idx in range(2, ws.max_row + 1):
    row_data = {}
    has_data = False
    
    for col_idx, header in enumerate(headers, start=1):
        cell_value = ws.cell(row_idx, col_idx).value
        if cell_value is not None:
            has_data = True
            if isinstance(cell_value, (int, float)):
                row_data[header] = cell_value
            else:
                row_data[header] = str(cell_value).strip()
    
    if has_data and len(row_data) > 0:
        fan_performance_db.append(row_data)

print(f"提取了 {len(fan_performance_db)} 条风机性能数据")

# ==================== 2. 提取风机相关标准 ====================
print("\n===== 提取风机相关标准 =====")
ws_standards_idx = wb.sheetnames[16]  # 索引16是"风机相关标准"
print(f"读取工作表: {ws_standards_idx}")
ws = wb[ws_standards_idx]

standards = []
std_headers = []

# 读取表头
for col_idx in range(1, min(10, ws.max_column + 1)):
    header = ws.cell(1, col_idx).value
    if header:
        std_headers.append(str(header).strip())

print(f"找到标准表头: {std_headers}")

# 读取标准数据
for row_idx in range(2, ws.max_row + 1):
    row_data = {}
    has_data = False
    
    for col_idx in range(1, min(10, ws.max_column + 1)):
        cell_value = ws.cell(row_idx, col_idx).value
        if cell_value is not None:
            has_data = True
            col_name = f"col{col_idx}"
            if col_idx - 1 < len(std_headers):
                col_name = std_headers[col_idx - 1]
            
            if isinstance(cell_value, (int, float)):
                row_data[col_name] = cell_value
            else:
                row_data[col_name] = str(cell_value).strip()
    
    if has_data and len(row_data) > 0:
        standards.append(row_data)

print(f"提取了 {len(standards)} 条风机标准数据")

# ==================== 3. 提取离心风机选型图表 ====================
print("\n===== 提取离心风机选型图表 =====")
lxfj_sheets = [idx for idx, name in enumerate(wb.sheetnames) if '离' in name and '风' in name and '选' in name]
print(f"找到离心风机选型相关工作表索引: {lxfj_sheets}")

calculation_data = {}
if len(lxfj_sheets) > 0:
    ws_calc_idx = wb.sheetnames[lxfj_sheets[0]]
    print(f"读取工作表: {ws_calc_idx}")
    ws = wb[ws_calc_idx]
    
    # 读取前30行前15列的所有数据
    for row_idx in range(1, min(31, ws.max_row + 1)):
        for col_idx in range(1, min(16, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            cell_value = cell.value
            
            if cell_value is not None:
                cell_key = f"R{row_idx}C{col_idx}"
                if isinstance(cell_value, (int, float)):
                    calculation_data[cell_key] = {"value": cell_value, "type": "number"}
                else:
                    cell_str = str(cell_value).strip()
                    calculation_data[cell_key] = {"value": cell_str, "type": "text"}
    
    print(f"提取了 {len(calculation_data)} 个单元格数据")

# ==================== 4. 保存为JSON ====================
output_data = {
    "fan_performance_db": fan_performance_db,
    "standards": standards,
    "calculation_data": calculation_data,
    "metadata": {
        "total_fan_models": len(fan_performance_db),
        "total_standards": len(standards),
        "source": "电机电力电气计算表（11月）.xlsx"
    }
}

with open('fan_system_full_data.json', 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print("\nJSON数据已保存到 fan_system_full_data.json")

# ==================== 5. 生成JavaScript数据文件 ====================
js_content = f"""// 风机选型系统完整数据库
// 自动生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

const FAN_PERFORMANCE_DB = {json.dumps(fan_performance_db, ensure_ascii=False, indent=2)};

const FAN_STANDARDS = {json.dumps(standards, ensure_ascii=False, indent=2)};

const FAN_CALCULATION_DATA = {json.dumps(calculation_data, ensure_ascii=False, indent=2)};

if (typeof module !== 'undefined' && module.exports) {{
  module.exports = {{ FAN_PERFORMANCE_DB, FAN_STANDARDS, FAN_CALCULATION_DATA }};
}}
"""

with open('dist-refactored/fan_system_database.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

print("JavaScript数据文件已保存到 dist-refactored/fan_system_database.js")

wb.close()
print("\n=== 数据提取完成 ===")
