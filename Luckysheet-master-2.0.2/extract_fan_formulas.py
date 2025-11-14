# -*- coding: utf-8 -*-
"""
提取风机选型计算表中的公式逻辑
"""

import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('电机电力电气计算表（11月）.xlsx')

# 获取"风机选型计算"工作表
ws = wb['风机选型计算 ']

print("=" * 80)
print("风机选型计算 - 关键单元格及其公式")
print("=" * 80)

# 定义要检查的关键单元格范围
key_cells = [
    ('C4', '流量'),
    ('C5', '全压'),
    ('C7', '工作温度'),
    ('C8', '绝热指数'),
    ('F4', '海拔高度'),
    ('F5', '当地大气压'),
    ('F6', '工况密度'),
    ('F7', '工作转速'),
    ('F8', '压缩性系数'),
    ('C9', '计算比转数'),
    ('F9', '风机型号'),
    ('C10', '粗算叶轮直径'),
    ('F10', '给定叶轮直径'),
]

print("\n【输入参数与计算结果】")
print("-" * 80)
for cell_addr, desc in key_cells:
    cell = ws[cell_addr]
    value = cell.value
    formula = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
    
    # 尝试获取公式（如果有）
    if hasattr(cell, 'data_type') and cell.data_type == 'f':
        formula = f"={cell.value}" if cell.value else None
    
    print(f"{cell_addr} ({desc}):")
    print(f"  值: {value}")
    if formula:
        print(f"  公式: {formula}")
    print()

# 扫描所有包含公式的单元格
print("\n【所有公式单元格】")
print("-" * 80)
formulas_found = {}

for row_idx in range(1, min(50, ws.max_row + 1)):
    for col_idx in range(1, min(15, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        
        # 检查是否是公式
        if cell.data_type == 'f':
            cell_addr = f"{chr(64+col_idx)}{row_idx}"
            formula = cell.value
            
            # 获取单元格左侧的标签（通常在B列）
            label_cell = ws.cell(row_idx, 2).value if col_idx > 2 else ws.cell(row_idx, 1).value
            label = str(label_cell).strip() if label_cell else ''
            
            formulas_found[cell_addr] = {
                'formula': formula,
                'label': label,
                'row': row_idx,
                'col': col_idx
            }
            
            print(f"{cell_addr} [{label}]: {formula}")

print(f"\n共找到 {len(formulas_found)} 个公式单元格")

# 保存公式到文件
import json
with open('fan_formulas.json', 'w', encoding='utf-8') as f:
    json.dump(formulas_found, f, ensure_ascii=False, indent=2)

print("\n公式已保存到 fan_formulas.json")

wb.close()
print("\n" + "=" * 80)


