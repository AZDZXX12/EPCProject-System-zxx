# -*- coding: utf-8 -*-
"""
最终正确提取YJV电缆数据
"""

import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('数据库.xlsx', data_only=True)
ws = wb['YJV']

print("开始提取YJV电缆数据...")

yjv_data = []

# 从第5行开始读取（前4行是表头）
for row_idx in range(5, ws.max_row + 1):
    spec = ws.cell(row_idx, 1).value  # A列：型号规格
    
    if not spec or spec == '':
        continue
    
    # 跳过非数据行
    spec_str = str(spec).strip()
    if '芯' in spec_str or 'YJV' in spec_str or spec_str == '':
        continue
    
    row_data = {
        "spec": spec_str,
        "insulation": ws.cell(row_idx, 2).value,  # B列：绝缘厚度
        "sheath": ws.cell(row_idx, 4).value,  # D列：护套厚度
        "approx_outer_diameter": ws.cell(row_idx, 5).value,  # E列：近似外径
        "approx_weight": ws.cell(row_idx, 6).value,  # F列：近似重量
        "max_conductor_dc_resistance": ws.cell(row_idx, 7).value,  # G列：直流电阻
        "test_voltage": ws.cell(row_idx, 8).value,  # H列：试验电压
        "air_current": ws.cell(row_idx, 9).value,  # I列：空气中载流量
        "soil_current": ws.cell(row_idx, 10).value  # J列：土壤中载流量
    }
    
    yjv_data.append(row_data)

print(f"共提取 {len(yjv_data)} 条YJV电缆数据")

# 打印前5条验证
print("\n前5条数据:")
for i, data in enumerate(yjv_data[:5], 1):
    print(f"{i}. {data['spec']}: 绝缘={data['insulation']}mm, 护套={data['sheath']}mm, "
          f"空气载流={data['air_current']}A, 土壤载流={data['soil_current']}A")

# 保存JSON
with open('yjv_full_table_correct.json', 'w', encoding='utf-8') as f:
    json.dump(yjv_data, f, ensure_ascii=False, indent=2)

# 生成JavaScript文件（给React使用）
js_react = f"""// YJV电缆完整数据表
// 提取自: 数据库.xlsx - YJV工作表
// 共 {len(yjv_data)} 条数据

const YJV_FULL_TABLE = {json.dumps(yjv_data, ensure_ascii=False, indent=2)};

if (typeof module !== 'undefined' && module.exports) {{
  module.exports = YJV_FULL_TABLE;
}}
"""

with open('dist-refactored/yjv_full_table.js', 'w', encoding='utf-8') as f:
    f.write(js_react)

# 同时生成TypeScript版本（给React TechnicalDataPanel使用）
ts_content = f"""// YJV电缆数据 - 用于React组件
export const YJV_CABLE_DATA = {json.dumps(yjv_data, ensure_ascii=False, indent=2)};
"""

with open('xuanxing/frontend/src/data/yjv_cable_data.ts', 'w', encoding='utf-8') as f:
    f.write(ts_content)

print("\n文件已生成:")
print("  - yjv_full_table_correct.json")
print("  - dist-refactored/yjv_full_table.js")
print("  - xuanxing/frontend/src/data/yjv_cable_data.ts")

wb.close()
print("\n✓ 提取完成！")


