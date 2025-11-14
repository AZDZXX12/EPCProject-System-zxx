# -*- coding: utf-8 -*-
"""
读取数据库.xlsx文件，查找离心风机系数表
"""

import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

print("读取 数据库.xlsx 文件...")
print("=" * 120)

try:
    wb = openpyxl.load_workbook('数据库.xlsx', data_only=True)
except FileNotFoundError:
    print("❌ 找不到 数据库.xlsx 文件！")
    exit(1)

print(f"✓ 成功打开文件\n")
print("工作表列表:")
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f"{idx}. {sheet_name} (行数: {ws.max_row}, 列数: {ws.max_column})")

# 查找包含风机系数的工作表
print("\n搜索包含'风机'、'型号'、'系数'的工作表...")
print("-" * 120)

fan_sheets = []
for sheet_name in wb.sheetnames:
    if '风' in sheet_name or 'fan' in sheet_name.lower() or '型' in sheet_name:
        fan_sheets.append(sheet_name)
        print(f"✓ 找到相关工作表: {sheet_name}")

# 逐个查看风机相关工作表
for sheet_name in fan_sheets:
    ws = wb[sheet_name]
    print(f"\n{'=' * 120}")
    print(f"工作表: {sheet_name}")
    print(f"{'=' * 120}")
    print(f"最大行数: {ws.max_row}, 最大列数: {ws.max_column}\n")
    
    # 显示前20行前20列
    print("前20行数据:")
    for row_idx in range(1, min(21, ws.max_row + 1)):
        print(f"\n第{row_idx}行:")
        for col_idx in range(1, min(21, ws.max_column + 1)):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                print(f"  {chr(64+col_idx) if col_idx <= 26 else 'A'+chr(38+col_idx)}{row_idx}: {val}")
    
    # 尝试提取数据
    print(f"\n提取 {sheet_name} 的数据...")
    
    # 假设第1行是表头
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        if header:
            headers.append(str(header).strip())
        else:
            headers.append(f"Col{col_idx}")
    
    print(f"表头: {headers[:15]}")
    
    # 提取数据
    data_rows = []
    for row_idx in range(2, min(100, ws.max_row + 1)):
        row_data = {}
        has_data = False
        
        for col_idx, header in enumerate(headers, start=1):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                has_data = True
                row_data[header] = val
        
        if has_data:
            data_rows.append(row_data)
    
    print(f"提取了 {len(data_rows)} 行数据")
    
    # 保存数据
    if len(data_rows) > 0:
        filename = f"database_{sheet_name}.json"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data_rows, f, ensure_ascii=False, indent=2)
        print(f"✓ 数据已保存到 {filename}")
        
        # 显示前2条数据示例
        print("\n前2条数据示例:")
        for i, row in enumerate(data_rows[:2], 1):
            print(f"\n  {i}. {list(row.items())[:10]}")

wb.close()
print("\n" + "=" * 120)
print("完成！")


