# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('电机电力电气计算表（11月）.xlsx', data_only=True)

# 搜索关键词
keywords = ['压力系数', '流量系数', '型号', '比转速', '效率']

print("搜索包含风机参数的工作表...\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # 检查前5行的所有列
    found_keywords = set()
    for row_idx in range(1, 6):
        for col_idx in range(1, min(30, ws.max_column + 1)):
            val = ws.cell(row_idx, col_idx).value
            if val:
                val_str = str(val)
                for kw in keywords:
                    if kw in val_str:
                        found_keywords.add(kw)
    
    if len(found_keywords) >= 2:  # 至少匹配2个关键词
        print(f"✓ {sheet_name}")
        print(f"  匹配关键词: {', '.join(found_keywords)}")
        print(f"  行数: {ws.max_row}, 列数: {ws.max_column}")
        
        # 显示前3行
        print("  前3行内容:")
        for row_idx in range(1, 4):
            print(f"    第{row_idx}行:", end="")
            for col_idx in range(1, min(15, ws.max_column + 1)):
                val = ws.cell(row_idx, col_idx).value
                if val:
                    print(f" [{chr(64+col_idx)}:{val}]", end="")
            print()
        print()

wb.close()


