import openpyxl
import json
import sys

# 确保输出使用UTF-8编码
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# 读取Excel文件
excel_file = '电机电力电气计算表（11月）.xlsx'
print(f"正在读取 {excel_file}...")

try:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # 初始化数据结构
    fan_data = {
        'performance': [],  # 风机性能表
        'standards': [],    # 风机相关标准
        'selection_charts': []  # 风机选型图表
    }
    
    # 1. 读取"风机性能表"
    if '风机性能表' in wb.sheetnames:
        print('\n正在读取风机性能表...')
        ws = wb['风机性能表']
        
        # 读取基本参数 (第2-7行)
        basic_params = {
            'impeller_diameter': ws.cell(2, 4).value,  # 叶轮直径
            'rated_speed': ws.cell(3, 4).value,        # 额定转速
            'air_temperature': ws.cell(4, 4).value,    # 空气温度
            'atmospheric_pressure': ws.cell(5, 4).value,  # 大气压力
            'air_density': ws.cell(6, 4).value         # 空气密度
        }
        
        # 读取性能数据 (从第9行开始)
        performance_data = []
        for row_idx in range(9, 15):  # 第9-14行是数据
            row = ws[row_idx]
            if row[0].value is not None:  # 如果序号不为空
                perf = {
                    'no': row[0].value,              # 序号
                    'pressure_coefficient': row[1].value,  # 压力系数
                    'flow_coefficient': row[2].value,      # 流量系数
                    'efficiency': row[3].value,            # 效率
                    'pressure': row[4].value,              # 压力
                    'flow': row[5].value,                  # 流量
                    'internal_power': row[6].value         # 内功率
                }
                performance_data.append(perf)
        
        fan_data['performance'] = {
            'basic_params': basic_params,
            'performance_data': performance_data
        }
        print(f'  - 读取了 {len(performance_data)} 条性能数据')
    
    # 2. 读取"风机相关标准"
    if '风机相关标准' in wb.sheetnames:
        print('正在读取风机相关标准...')
        ws = wb['风机相关标准']
        
        # 从第3行开始读取标准 (第1-2行是标题)
        for row_idx in range(3, ws.max_row + 1):
            row = ws[row_idx]
            if row[0].value is not None:  # 如果序号不为空
                standard = {
                    'no': row[0].value,
                    'standard_code': row[1].value,
                    'standard_name': row[2].value,
                    'application_scope': row[3].value if len(row) > 3 else None
                }
                fan_data['standards'].append(standard)
        
        print(f'  - 读取了 {len(fan_data["standards"])} 条标准')
    
    # 3. 读取"离风选型图表1"
    if '离风选型图表1' in wb.sheetnames:
        print('正在读取离风选型图表1...')
        ws = wb['离风选型图表1']
        
        # 读取关键参数 (前几行)
        selection_params = {
            'air_volume': ws.cell(4, 3).value,           # 风量
            'total_pressure': ws.cell(5, 3).value,       # 全压
            'static_pressure': ws.cell(6, 3).value,      # 静压
            'air_temperature': ws.cell(7, 3).value,      # 空气温度
            'pressure_index': ws.cell(8, 3).value,       # 压缩指数
            'altitude': ws.cell(4, 6).value,             # 海拔高度
            'atmospheric_pressure': ws.cell(5, 7).value, # 大气压力
            'air_density': ws.cell(6, 7).value,          # 空气密度
            'fan_speed': ws.cell(7, 7).value             # 风机转速
        }
        
        fan_data['selection_params'] = selection_params
        print('  - 读取了选型参数')
    
    # 保存为JSON文件
    output_file = 'dist-refactored/fan_data.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(fan_data, f, ensure_ascii=False, indent=2)
    print(f'\n成功生成 {output_file}')
    
    # 生成JavaScript文件
    js_file = 'dist-refactored/fan_database.js'
    with open(js_file, 'w', encoding='utf-8') as f:
        f.write('// 风机数据库 - 从 电机电力电气计算表（11月）.xlsx 自动生成\n')
        f.write('// 生成时间: 自动生成\n\n')
        f.write('const FAN_DATABASE = ')
        f.write(json.dumps(fan_data, ensure_ascii=False, indent=2))
        f.write(';\n\n')
        f.write('// 导出数据\n')
        f.write('if (typeof module !== "undefined" && module.exports) {\n')
        f.write('  module.exports = FAN_DATABASE;\n')
        f.write('}\n')
    
    print(f'成功生成 {js_file}')
    print('\n数据统计:')
    print(f'  - 性能数据: {len(fan_data["performance"].get("performance_data", []))} 条')
    print(f'  - 标准数据: {len(fan_data["standards"])} 条')
    
except FileNotFoundError:
    print(f'错误: 找不到文件 {excel_file}')
except Exception as e:
    print(f'错误: {e}')
    import traceback
    traceback.print_exc()
