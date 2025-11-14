#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
深度分析风机选型Excel表格的计算逻辑和公式
"""
import openpyxl
import sys

# 设置UTF-8输出
sys.stdout.reconfigure(encoding='utf-8')

# 打开Excel文件
workbook = openpyxl.load_workbook('电机电力电气计算表（11月）.xlsx', data_only=False)

# 获取"风机选型计算"工作表
if '风机选型计算' in workbook.sheetnames:
    calc_sheet = workbook['风机选型计算']
    print("=" * 80)
    print("【风机选型计算】工作表分析")
    print("=" * 80)
    
    # 分析前20行的所有列（A-Z）
    print("\n关键单元格及其公式：")
    print("-" * 80)
    
    # 重点分析C列到G列的前30行
    for row in range(1, 31):
        row_data = []
        has_content = False
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            cell = calc_sheet[f'{col}{row}']
            value = cell.value
            
            if value is not None:
                has_content = True
                
                # 如果是公式，显示公式
                if isinstance(value, str) and value.startswith('='):
                    row_data.append(f"{col}{row}: 公式={value}")
                else:
                    row_data.append(f"{col}{row}: {value}")
        
        if has_content:
            print(f"\n第{row}行:")
            for item in row_data:
                print(f"  {item}")
    
    print("\n" + "=" * 80)
    print("特别关注的计算单元格（C10-G20）：")
    print("=" * 80)
    
    for row in range(10, 21):
        for col in ['C', 'D', 'E', 'F', 'G']:
            cell = calc_sheet[f'{col}{row}']
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"\n{col}{row} = {cell.value}")
                    # 尝试获取计算后的值
                    calc_wb = openpyxl.load_workbook('电机电力电气计算表（11月）.xlsx', data_only=True)
                    calc_value = calc_wb['风机选型计算'][f'{col}{row}'].value
                    print(f"  → 计算结果: {calc_value}")
                else:
                    print(f"{col}{row} = {cell.value}")

# 获取"风机性能表"工作表 - 查看型号列表和数据结构
if '风机性能表' in workbook.sheetnames:
    perf_sheet = workbook['风机性能表']
    print("\n" + "=" * 80)
    print("【风机性能表】工作表结构分析")
    print("=" * 80)
    
    # 读取前10行以了解结构
    print("\n表头结构（前3行，A-M列）:")
    for row in range(1, 4):
        row_data = []
        for col_idx in range(1, 14):  # A-M
            cell = perf_sheet.cell(row=row, column=col_idx)
            if cell.value:
                row_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}{row}:{cell.value}")
        if row_data:
            print(f"  第{row}行: {' | '.join(row_data)}")
    
    # 查看风机型号列（通常在第一列）
    print("\n风机型号列表（前20个）:")
    for row in range(4, 24):
        model_cell = perf_sheet.cell(row=row, column=1)
        if model_cell.value:
            print(f"  行{row}: {model_cell.value}")

print("\n" + "=" * 80)
print("分析完成！")
print("=" * 80)


