# -*- coding: utf-8 -*-
"""
正确提取YJV电缆数据表 - 确保列名对应正确
"""

import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 打开Excel文件
wb = openpyxl.load_workbook('数据库.xlsx', data_only=True)

# 获取YJV工作表
ws = wb['YJV']

print(f"工作表名称: YJV")
print(f"最大行数: {ws.max_row}, 最大列数: {ws.max_column}")

# 首先打印表头，确认列名
print("\n表头内容:")
for col_idx in range(1, min(15, ws.max_column + 1)):
    header = ws.cell(1, col_idx).value
    print(f"列{col_idx} ({chr(64+col_idx)}): {header}")

# 根据您的截图，正确的列顺序应该是：
# A: 型号规格(mm²)
# B: 绝缘层厚度(mm)
# C: 护套层厚度(mm)
# D: 电缆近似外径(mm)
# E: 电缆近似重量(kg/km)
# F: 导体最大直流电阻(Ω/km)
# G: 试验电压(KV)
# H: 在空气中载流量(A)
# I: 直埋土壤中载流量(A)

yjv_data = []

for row_idx in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
    spec = ws.cell(row_idx, 1).value  # A列：型号规格
    
    if not spec:  # 如果型号为空，跳过
        continue
    
    row_data = {
        "spec": str(spec).strip(),
        "insulation": ws.cell(row_idx, 2).value,  # B列：绝缘层厚度
        "sheath": ws.cell(row_idx, 3).value,  # C列：护套层厚度
        "approx_outer_diameter": ws.cell(row_idx, 4).value,  # D列：近似外径
        "approx_weight": ws.cell(row_idx, 5).value,  # E列：近似重量
        "max_conductor_dc_resistance": ws.cell(row_idx, 6).value,  # F列：直流电阻
        "test_voltage": ws.cell(row_idx, 7).value,  # G列：试验电压
        "air_current": ws.cell(row_idx, 8).value,  # H列：空气中载流量
        "soil_current": ws.cell(row_idx, 9).value  # I列：土壤中载流量
    }
    
    yjv_data.append(row_data)

print(f"\n共提取 {len(yjv_data)} 条YJV电缆数据")

# 打印前3条数据验证
print("\n前3条数据示例:")
for i, data in enumerate(yjv_data[:3], 1):
    print(f"{i}. {data['spec']}: 空气载流量={data['air_current']}A, 土壤载流量={data['soil_current']}A")

# 保存为JSON
with open('yjv_full_table_correct.json', 'w', encoding='utf-8') as f:
    json.dump(yjv_data, f, ensure_ascii=False, indent=2)

# 生成JavaScript文件
js_content = f"""// YJV电缆完整数据表 - 从 数据库.xlsx 正确提取
// 自动生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

const YJV_FULL_TABLE = {json.dumps(yjv_data, ensure_ascii=False, indent=2)};

if (typeof module !== 'undefined' && module.exports) {{
  module.exports = YJV_FULL_TABLE;
}}
"""

with open('dist-refactored/yjv_full_table.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

print(f"\nJSON文件已保存: yjv_full_table_correct.json")
print(f"JavaScript文件已保存: dist-refactored/yjv_full_table.js")

wb.close()
print("\n提取完成！")


