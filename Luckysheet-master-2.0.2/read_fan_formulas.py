import openpyxl
import json
import sys

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

excel_file = '电机电力电气计算表（11月）.xlsx'
print(f"读取 {excel_file} 中的风机选型计算公式...\n")

try:
    wb = openpyxl.load_workbook(excel_file, data_only=False)  # 读取公式
    
    if '风机选型计算' in wb.sheetnames:
        ws = wb['风机选型计算']
        
        formulas = {}
        
        # 读取关键单元格的公式
        print("=" * 80)
        print("关键计算公式：")
        print("=" * 80)
        
        # 第4行：流量、全压、进口压力、工作温度
        formulas['airflow'] = {'cell': 'C4', 'value': ws['C4'].value, 'formula': ws['C4'].value}
        formulas['total_pressure'] = {'cell': 'C5', 'value': ws['C5'].value, 'formula': ws['C5'].value}
        formulas['inlet_pressure'] = {'cell': 'C6', 'value': ws['C6'].value, 'formula': ws['C6'].value}
        formulas['temperature'] = {'cell': 'C7', 'value': ws['C7'].value, 'formula': ws['C7'].value}
        formulas['adiabatic_index'] = {'cell': 'C8', 'value': ws['C8'].value, 'formula': ws['C8'].value}
        
        # 第4-10行：关键计算公式
        formulas['altitude'] = {'cell': 'F4', 'value': ws['F4'].value}
        formulas['atmospheric_pressure'] = {'cell': 'F5', 'value': ws['F5'].value, 'formula': ws['F5'].value}
        formulas['working_density'] = {'cell': 'F6', 'value': ws['F6'].value, 'formula': ws['F6'].value}
        formulas['working_speed'] = {'cell': 'F7', 'value': ws['F7'].value, 'formula': ws['F7'].value}
        formulas['compression_coeff'] = {'cell': 'F8', 'value': ws['F8'].value, 'formula': ws['F8'].value}
        formulas['specific_speed'] = {'cell': 'F9', 'value': ws['F9'].value, 'formula': ws['F9'].value}
        formulas['impeller_diameter'] = {'cell': 'F10', 'value': ws['F10'].value, 'formula': ws['F10'].value}
        
        # 读取前20行所有内容
        print("\n前20行数据：")
        for i in range(1, 21):
            row_data = []
            for j in range(1, 11):
                cell = ws.cell(i, j)
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        row_data.append(f"公式: {cell.value}")
                    else:
                        row_data.append(str(cell.value))
            if row_data:
                print(f"第{i:2d}行: {row_data}")
        
        # 保存公式到JSON
        with open('dist-refactored/fan_formulas.json', 'w', encoding='utf-8') as f:
            json.dump(formulas, f, ensure_ascii=False, indent=2)
        
        print("\n公式已保存到 fan_formulas.json")
        
except Exception as e:
    print(f'错误: {e}')
    import traceback
    traceback.print_exc()


