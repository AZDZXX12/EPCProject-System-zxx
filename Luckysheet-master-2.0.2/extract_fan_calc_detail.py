# -*- coding: utf-8 -*-
"""
详细提取"风机选型计算"工作表数据
"""

import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('电机电力电气计算表（11月）.xlsx', data_only=True)

# 获取"风机选型计算 "工作表（注意有空格）
ws_name = '风机选型计算 '
ws = wb[ws_name]

print(f"读取工作表: {ws_name}")
print(f"最大行数: {ws.max_row}, 最大列数: {ws.max_column}")

# 提取前40行前20列的所有内容
calc_matrix = []
for row_idx in range(1, min(41, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(21, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        row_data.append({
            "coord": f"{chr(64+col_idx)}{row_idx}",
            "value": cell.value,
            "type": type(cell.value).__name__
        })
    calc_matrix.append(row_data)

# 保存详细数据
with open('fan_calc_detail.json', 'w', encoding='utf-8') as f:
    json.dump(calc_matrix, f, ensure_ascii=False, indent=2)

print("详细数据已保存到 fan_calc_detail.json")

# 打印前10行的关键单元格
print("\n前10行关键单元格:")
for row_idx in range(1, 11):
    for col_idx in range(1, 11):
        cell = ws.cell(row_idx, col_idx)
        if cell.value:
            print(f"  {chr(64+col_idx)}{row_idx}: {cell.value}")

wb.close()


