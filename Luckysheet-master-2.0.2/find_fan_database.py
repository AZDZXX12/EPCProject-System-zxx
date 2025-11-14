# -*- coding: utf-8 -*-
"""
查找包含风机型号和系数的工作表
根据公式 VLOOKUP(F9,[2]数据库!A$1:S$65536,14)
需要找到包含风机型号列表的表
"""

import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('电机电力电气计算表（11月）.xlsx', data_only=True)

print("搜索包含风机型号的工作表...")
print("=" * 100)

# 已知的风机型号示例（从公式中看到F9='5-48'等）
known_models = ['5-48', 'BB24', 'BB50', '4-72', '4-68', '4-79']

for sheet_idx, sheet_name in enumerate(wb.sheetnames):
    ws = wb[sheet_name]
    
    # 检查第一列是否包含风机型号
    found_models = []
    for row_idx in range(1, min(100, ws.max_row + 1)):
        cell_val = ws.cell(row_idx, 1).value
        if cell_val:
            cell_str = str(cell_val).strip()
            for model in known_models:
                if model in cell_str:
                    found_models.append(cell_str)
                    break
    
    if found_models:
        print(f"\n✓ 工作表 #{sheet_idx+1}: {sheet_name}")
        print(f"  找到风机型号: {found_models[:5]}")
        
        # 显示此表的结构
        print(f"  最大行数: {ws.max_row}, 最大列数: {ws.max_column}")
        
        # 显示表头
        print("  表头:")
        for col_idx in range(1, min(20, ws.max_column + 1)):
            header = ws.cell(1, col_idx).value
            if header:
                print(f"    列{col_idx}({chr(64+col_idx)}): {header}")
        
        # 显示第一条数据
        print("  第一条数据示例:")
        for col_idx in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(2, col_idx).value
            if val:
                print(f"    列{col_idx}: {val}")

wb.close()
print("\n" + "=" * 100)


