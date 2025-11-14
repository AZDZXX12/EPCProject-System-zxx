# -*- coding: utf-8 -*-
"""
提取[2]数据库工作表中的离心风机系数表
这是Excel公式中VLOOKUP引用的关键数据源
"""

import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('电机电力电气计算表（11月）.xlsx', data_only=True)

# 查找包含"数据库"的工作表
print("工作表列表:")
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"{idx}. {sheet_name}")
    if '数据库' in sheet_name or 'database' in sheet_name.lower():
        print(f"   ^^^ 找到数据库工作表！")

# 尝试不同的可能名称
possible_names = ['[2]数据库', '数据库', 'database', 'Database']
ws = None

for name in possible_names:
    try:
        ws = wb[name]
        print(f"\n成功打开工作表: {name}")
        break
    except:
        continue

if ws is None:
    # 尝试索引方式（如果[2]表示第2个工作表）
    try:
        ws = wb.worksheets[1]  # 索引1是第2个工作表
        print(f"\n使用索引打开第2个工作表: {ws.title}")
    except:
        print("错误：找不到数据库工作表！")
        wb.close()
        exit(1)

print(f"工作表名称: {ws.title}")
print(f"最大行数: {ws.max_row}, 最大列数: {ws.max_column}")

# 打印前10行的所有列，查看结构
print("\n【前10行数据结构】")
print("=" * 120)
for row_idx in range(1, min(11, ws.max_row + 1)):
    print(f"\n第{row_idx}行:")
    for col_idx in range(1, min(30, ws.max_column + 1)):  # 读取前30列
        val = ws.cell(row_idx, col_idx).value
        if val is not None:
            print(f"  {chr(64+col_idx)}{row_idx}: {val}")

# 读取表头
print("\n【提取风机系数表】")
print("=" * 120)
headers = []
header_row = 1  # 假设第1行是表头

for col_idx in range(1, min(30, ws.max_column + 1)):
    header = ws.cell(header_row, col_idx).value
    if header:
        headers.append(str(header).strip())
    else:
        headers.append(f"Col{col_idx}")

print(f"表头: {headers[:20]}")  # 显示前20列

# 提取风机数据
fan_coefficients = []

for row_idx in range(2, min(50, ws.max_row + 1)):  # 读取前50行数据
    row_data = {}
    has_data = False
    
    # 读取第1列作为风机型号
    fan_model = ws.cell(row_idx, 1).value
    
    if not fan_model:
        continue
    
    row_data['风机型号'] = str(fan_model).strip()
    has_data = True
    
    # 读取所有列的数据
    for col_idx in range(2, min(30, ws.max_column + 1)):
        val = ws.cell(row_idx, col_idx).value
        if val is not None:
            col_name = headers[col_idx-1] if col_idx-1 < len(headers) else f"Col{col_idx}"
            row_data[col_name] = val
    
    if has_data:
        fan_coefficients.append(row_data)

print(f"\n共提取 {len(fan_coefficients)} 种风机型号的系数数据")

# 显示前3条数据
print("\n【前3条风机数据示例】")
for i, data in enumerate(fan_coefficients[:3], 1):
    print(f"\n{i}. 风机型号: {data.get('风机型号', 'N/A')}")
    for key, val in list(data.items())[:10]:  # 显示前10个字段
        if key != '风机型号':
            print(f"   {key}: {val}")

# 保存为JSON
with open('fan_coefficients.json', 'w', encoding='utf-8') as f:
    json.dump(fan_coefficients, f, ensure_ascii=False, indent=2)

# 生成JavaScript文件
js_content = f"""// 离心风机系数数据库
// 提取自: 电机电力电气计算表（11月）.xlsx - [2]数据库工作表
// 共 {len(fan_coefficients)} 种风机型号

const FAN_COEFFICIENTS_DB = {json.dumps(fan_coefficients, ensure_ascii=False, indent=2)};

// 根据风机型号查找系数
function getFanCoefficient(fanModel, columnIndex) {{
    const fan = FAN_COEFFICIENTS_DB.find(f => f['风机型号'] === fanModel);
    if (!fan) return null;
    
    // 返回指定列的数据
    const keys = Object.keys(fan);
    return fan[keys[columnIndex]] || null;
}}

if (typeof module !== 'undefined' && module.exports) {{
    module.exports = {{ FAN_COEFFICIENTS_DB, getFanCoefficient }};
}}
"""

with open('dist-refactored/fan_coefficients.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

print(f"\n✓ JSON已保存: fan_coefficients.json")
print(f"✓ JavaScript已保存: dist-refactored/fan_coefficients.js")

wb.close()
print("\n" + "=" * 120)


