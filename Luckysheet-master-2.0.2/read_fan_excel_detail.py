import openpyxl
import sys

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

excel_file = '电机电力电气计算表（11月）.xlsx'
print(f"正在详细读取 {excel_file}...\n")

try:
    wb = openpyxl.load_workbook(excel_file, data_only=False)  # 读取公式
    
    # 读取"离风选型图表1"工作表
    if '离风选型图表1' in wb.sheetnames:
        print('='*80)
        print('工作表: 离风选型图表1 (详细读取)')
        print('='*80)
        
        ws = wb['离风选型图表1']
        
        # 读取前50行，包括公式
        for i in range(1, min(51, ws.max_row + 1)):
            row_data = []
            for j in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(i, j)
                if cell.value is not None:
                    if cell.data_type == 'f':  # 公式
                        row_data.append(f"[公式: {cell.value}]")
                    else:
                        row_data.append(str(cell.value))
                else:
                    row_data.append('')
            
            if any(row_data):  # 只打印非空行
                print(f'第{i:3d}行: {row_data}')
        
        print(f'\n总行数: {ws.max_row}')
        print(f'总列数: {ws.max_column}')
    
    # 读取"风机选型计算"相关工作表
    print('\n\n' + '='*80)
    print('查找包含"选型"或"计算"的工作表...')
    print('='*80)
    
    for sheet_name in wb.sheetnames:
        if '选型' in sheet_name or '计算' in sheet_name:
            print(f'\n找到: {sheet_name}')
            ws = wb[sheet_name]
            print(f'  行数: {ws.max_row}, 列数: {ws.max_column}')
            
            # 读取前20行
            print('  前20行内容:')
            for i in range(1, min(21, ws.max_row + 1)):
                row_data = []
                for j in range(1, min(11, ws.max_column + 1)):
                    cell = ws.cell(i, j)
                    if cell.value is not None:
                        if cell.data_type == 'f':
                            row_data.append(f"[={cell.value}]")
                        else:
                            row_data.append(str(cell.value))
                
                if row_data:
                    print(f'    {i}: {row_data}')

except Exception as e:
    print(f'错误: {e}')
    import traceback
    traceback.print_exc()


