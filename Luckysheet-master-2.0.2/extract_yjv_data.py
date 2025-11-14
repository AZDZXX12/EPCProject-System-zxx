import openpyxl
import json

# 读取Excel文件
print("正在读取数据库.xlsx...")
wb = openpyxl.load_workbook('数据库.xlsx', data_only=True)
print(f'工作表列表: {wb.sheetnames}\n')

# 检查是否有YJV工作表
if 'YJV' in wb.sheetnames:
    ws = wb['YJV']
    print('=' * 80)
    print('YJV工作表数据：')
    print('=' * 80)
    
    # 读取所有数据
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            print(f'表头: {row}')
            print('-' * 80)
        if i < 30:  # 显示前30行
            print(f'第{i+1:3d}行: {row}')
        data.append(list(row))
    
    # 保存为JSON
    json_file = 'dist-refactored/yjv_database.json'
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print('=' * 80)
    print(f'OK Saved {len(data)} rows to {json_file}')
    
    # 生成JavaScript数据格式
    js_file = 'dist-refactored/yjv_database.js'
    with open(js_file, 'w', encoding='utf-8') as f:
        f.write('// YJV电缆数据库 - 从Excel自动生成\n')
        f.write('// 数据源: 数据库.xlsx - YJV工作表\n\n')
        f.write('const YJV_DATABASE = ')
        f.write(json.dumps(data, ensure_ascii=False, indent=2))
        f.write(';\n\n')
        f.write('// 导出供页面使用\n')
        f.write('if (typeof window !== "undefined") {\n')
        f.write('    window.YJV_DATABASE = YJV_DATABASE;\n')
        f.write('}\n')
    
    print(f'OK Generated JavaScript file: {js_file}')
    
else:
    print('❌ 未找到YJV工作表')
    print(f'可用工作表: {wb.sheetnames}')

